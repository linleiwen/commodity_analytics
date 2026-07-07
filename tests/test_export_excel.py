"""Excel + QA exporter tests: verify all tabs are produced from a minimal DB."""

from __future__ import annotations

from openpyxl import load_workbook

EXPECTED_TABS = [
    "Ranked_Products", "Product_Master", "Price_Observations", "Demand_Signals",
    "Compliance_QA", "Luggage_Plan", "Assumptions", "Manual_Review", "Source_Log",
]


def _seed_minimal_db():
    from app.storage import db

    db.init_db()
    with db.connect() as conn:
        db.ensure_run(conn, "run")
        db.set_run_assumptions(conn, "run", '{"jpy_to_usd_rate": 0.0067, "fees": {"platform_fee_pct": 0.1325}}')
        db.upsert(conn, "product_master", {
            "product_id": "p", "canonical_name_en": "Test Gift", "category": "non_food_gift",
            "package_volume_liter": 0.5, "package_weight_g": 100, "melt_risk_level": "LOW",
            "crush_risk_level": "LOW", "leak_risk_level": "LOW", "risk_notes": "n/a",
        })
        db.upsert(conn, "price_observations", {
            "observation_id": "o1", "run_id": "run", "product_id": "p", "country": "JP",
            "platform": "FieldSurvey", "source_type": "manual", "currency": "JPY",
            "price": 500, "price_usd": 3.35, "match_confidence": 0.95,
            "confidence_level": "manual_verified",
        })
        db.upsert(conn, "demand_signals", {
            "signal_id": "d1", "run_id": "run", "product_id": "p", "signal_kind": "sell_through",
            "source_name": "eBayTerapeak", "median_sold_price": 20.0, "sold_count": 50,
        })
        db.upsert(conn, "compliance_reviews", {
            "review_id": "c", "run_id": "run", "product_id": "p", "compliance_status": "GREEN",
            "category_risk": "GREEN", "shipping_risk": "GREEN", "food_class": "",
            "reason_codes_json": "[]", "review_status": "auto",
        })
        db.upsert(conn, "scores", {
            "score_id": "s", "run_id": "run", "product_id": "p", "final_score": 80,
            "priority_tier": "A", "expected_net_profit_usd": 5.0, "recommended_qty": 2,
            "volume_used_liter": 1.0, "weight_used_lb": 0.44, "estimated_total_profit": 10.0,
            "reason_codes_json": "[]", "data_confidence": 0.85, "jp_cost_usd": 3.35,
            "us_expected_sold_price_usd": 20.0, "recommended_channel": "eBay + local pickup",
            "packing_notes": "standard packing",
        })
        db.log_source(conn, {
            "log_id": "l1", "run_id": "run", "source_name": "FieldSurvey",
            "collector": "manual_field_prices", "status": "success", "records": 1,
        })


def test_export_workbook_has_all_tabs(tmp_path, monkeypatch):
    from app import settings
    from app.exporters import excel as excel_exporter
    from app.exporters import qa_report

    monkeypatch.setattr(settings, "DB_PATH", tmp_path / "t.sqlite")
    _seed_minimal_db()

    out = tmp_path / "wb.xlsx"
    excel_exporter.export_workbook("run", out)
    assert out.exists()

    wb = load_workbook(out)
    for tab in EXPECTED_TABS:
        assert tab in wb.sheetnames

    ws = wb["Ranked_Products"]
    assert ws["A1"].value == "Rank"
    assert ws.freeze_panes == "A2"

    qa_path = tmp_path / "qa.md"
    qa_report.export_report("run", qa_path)
    text = qa_path.read_text(encoding="utf-8")
    assert "QA Report" in text and "run" in text
