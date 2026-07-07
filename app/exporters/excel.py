"""Excel workbook exporter (spec section 11).

Builds ``japan_dmv_product_rankings.xlsx`` with all nine tabs, conditional formatting on
the decision tab, frozen headers, and auto-filters. Uses openpyxl only.
"""

from __future__ import annotations

import json
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.exporters import dataset
from app.storage import db

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TIER_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="DDEBF7"),
    "C": PatternFill("solid", fgColor="E2EFDA"),
    "Watchlist": PatternFill("solid", fgColor="FFEB9C"),
    "Blocked": PatternFill("solid", fgColor="D9D9D9"),
}
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT = Font(color="9C0006")
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")

_MONEY_COLS = {"JP_Buy_Price_USD", "US_Expected_Sold_Price_USD", "Expected_Net_Profit_USD",
               "Profit_Per_Liter", "Profit_Per_Lb", "Estimated_Total_Profit",
               "Unit_Volume_Liter", "Total_Volume_Liter", "Unit_Weight_Lb", "Total_Weight_Lb"}


def _clean(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if pd.isna(value) if not isinstance(value, (list, dict)) else False:
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            return value
    return value


def _write_sheet(wb: Workbook, title: str, df: pd.DataFrame, freeze: bool = True) -> None:
    ws = wb.create_sheet(title=title[:31])
    if df is None or df.empty:
        ws["A1"] = "(no data)"
        return
    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([_clean(v) for v in row.tolist()])

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        letter = get_column_letter(col_idx)
        max_len = max([len(str(name))] + [len(str(_clean(v) or "")) for v in df[name].tolist()[:200]])
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 46)
        if name in _MONEY_COLS:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "#,##0.00"

    if freeze:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _format_ranked(ws, headers: list[str]) -> None:
    idx = {name: i + 1 for i, name in enumerate(headers)}
    tier_c = idx.get("Priority_Tier")
    comp_c = idx.get("Compliance_Status")
    shelf_c = idx.get("Shelf_Life_Days_Remaining")
    conf_c = idx.get("Data_Confidence")
    for r in range(2, ws.max_row + 1):
        if tier_c:
            tier = ws.cell(row=r, column=tier_c).value
            fill = _TIER_FILL.get(str(tier))
            if fill:
                ws.cell(row=r, column=tier_c).fill = fill
                ws.cell(row=r, column=tier_c).font = Font(bold=True)
        if comp_c and str(ws.cell(row=r, column=comp_c).value) == "RED":
            c = ws.cell(row=r, column=comp_c)
            c.fill, c.font = _RED_FILL, _RED_FONT
        if shelf_c:
            v = ws.cell(row=r, column=shelf_c).value
            if isinstance(v, (int, float)) and v < 90:
                ws.cell(row=r, column=shelf_c).fill = _YELLOW_FILL
        if conf_c:
            v = ws.cell(row=r, column=conf_c).value
            if isinstance(v, (int, float)) and v < 0.75:
                ws.cell(row=r, column=conf_c).fill = _ORANGE_FILL


def _luggage_plan(run_id: str) -> pd.DataFrame:
    scores = db.read_table("scores", run_id)
    products = db.read_products()
    if scores.empty:
        return pd.DataFrame()
    buy = scores[scores["recommended_qty"] > 0].copy()
    if buy.empty:
        return pd.DataFrame()
    name = dict(zip(products["product_id"], products["canonical_name_en"]))
    buy["Product"] = buy["product_id"].map(name)
    buy["Unit_Volume_Liter"] = (buy["volume_used_liter"] / buy["recommended_qty"]).round(3)
    buy["Unit_Weight_Lb"] = (buy["weight_used_lb"] / buy["recommended_qty"]).round(3)
    out = pd.DataFrame({
        "Product": buy["Product"],
        "Recommended_Qty": buy["recommended_qty"],
        "Unit_Volume_Liter": buy["Unit_Volume_Liter"],
        "Total_Volume_Liter": buy["volume_used_liter"].round(3),
        "Unit_Weight_Lb": buy["Unit_Weight_Lb"],
        "Total_Weight_Lb": buy["weight_used_lb"].round(3),
        "Estimated_Total_Profit": buy["estimated_total_profit"].round(2),
        "Packing_Instructions": buy["packing_notes"],
    })
    return out.sort_values("Estimated_Total_Profit", ascending=False).reset_index(drop=True)


def _assumptions_df(run_id: str) -> pd.DataFrame:
    row = db.read_df("SELECT assumptions_json FROM runs WHERE run_id = ?", (run_id,))
    if row.empty or not row.iloc[0]["assumptions_json"]:
        return pd.DataFrame({"Assumption": [], "Value": []})
    data = json.loads(row.iloc[0]["assumptions_json"])
    flat: list[tuple[str, str]] = []

    def _walk(prefix: str, obj) -> None:
        if isinstance(obj, dict):
            for k, v in obj.items():
                _walk(f"{prefix}.{k}" if prefix else k, v)
        else:
            flat.append((prefix, str(obj)))

    _walk("", data)
    return pd.DataFrame(flat, columns=["Assumption", "Value"])


def _select(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return df
    keep = [c for c in cols if c in df.columns]
    return df[keep]


def export_workbook(run_id: str, out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    ranked = dataset.build_ranked(run_id)
    _write_sheet(wb, "Ranked_Products", ranked)
    if "Ranked_Products" in wb.sheetnames and not ranked.empty:
        _format_ranked(wb["Ranked_Products"], list(ranked.columns))

    products = db.read_products()
    _write_sheet(wb, "Product_Master", _select(products, [
        "product_id", "canonical_name_en", "canonical_name_jp", "canonical_name_cn", "brand",
        "category", "subcategory", "jan_gtin", "package_count", "unit_size_text",
        "package_weight_g", "package_volume_liter", "is_food", "is_cosmetic",
        "is_drug_or_otc_risk", "is_hazmat_shipping_risk", "storage_condition",
        "melt_risk_level", "crush_risk_level", "leak_risk_level", "risk_notes",
    ]))

    obs = db.read_table("price_observations", run_id)
    _write_sheet(wb, "Price_Observations", _select(obs, [
        "product_id", "source_name", "source_type", "country", "platform", "listing_title",
        "currency", "price", "price_usd", "shipping_price", "availability_status",
        "expiration_date_parsed", "match_confidence", "confidence_level", "source_url",
    ]))

    sig = db.read_table("demand_signals", run_id)
    _write_sheet(wb, "Demand_Signals", _select(sig, [
        "product_id", "source_name", "signal_kind", "query", "window_days", "mention_count",
        "post_count", "like_count", "comment_count", "buyer_intent_count", "manual_heat_label",
        "sold_count", "sell_through_pct", "median_sold_price", "competitor_count",
        "days_to_sell", "trend_slope", "confidence_level", "notes",
    ]))

    comp = db.read_table("compliance_reviews", run_id)
    _write_sheet(wb, "Compliance_QA", _select(comp, [
        "product_id", "compliance_status", "category_risk", "food_safety_risk",
        "cosmetic_drug_risk", "shipping_risk", "labeling_risk", "food_class",
        "platform_risk_ebay", "platform_risk_facebook", "platform_risk_nextdoor",
        "platform_risk_xhs", "review_status", "reason_codes_json",
    ]))

    _write_sheet(wb, "Luggage_Plan", _luggage_plan(run_id))
    _write_sheet(wb, "Assumptions", _assumptions_df(run_id), freeze=False)
    _write_sheet(wb, "Manual_Review", dataset.build_manual_review(run_id))
    _write_sheet(wb, "Source_Log", _select(db.read_table("source_log", run_id), [
        "source_name", "collector", "query", "status", "http_status", "records",
        "message", "observed_at", "url", "snapshot_path",
    ]))

    wb.save(out)
    return out
